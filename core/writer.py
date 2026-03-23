"""
core/writer.py — Excel workbook / sheet writing logic.
B&W print optimized — compact horizontal, tall vertical, repeat headers.
All sheets are protected with password 507 (view-only, no edits).
"""

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import (
    CENTER_TITLE, SECTOR_SHEETS, NOTES_MERGE_HEADER,
    HEADER_FILL, ALT_FILL, WHITE_FILL, ODD_FILL,
    EMPTY_CELL_FILL,
    HDR_FONT, SUM_FONT, DATA_FONT, DATE_FONT,
    BORDER, THICK_BORDER,
    C_BLACK,
)

SHEET_PASSWORD = "507"


# ── Low-level helpers ─────────────────────────────────────────────────────────

def _al(h: str = "center", v: str = "center",
        ro: int = 2, wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, readingOrder=ro, wrap_text=wrap)


def _protect(ws: Worksheet) -> None:
    ws.protection.sheet    = True
    ws.protection.password = SHEET_PASSWORD
    ws.protection.enable()


def _print_setup(ws: Worksheet, row_header_end: int = 3) -> None:
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left   = ws.page_margins.right  = 0.30
    ws.page_margins.top    = ws.page_margins.bottom = 0.30
    ws.page_margins.header = ws.page_margins.footer = 0.15
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = f"1:{row_header_end}"


def _auto_col_widths(ws: Worksheet, num_cols: int,
                     num_data_rows: int, dispatch_col: int = 4) -> None:
    TOTAL_WIDTH = 80.0
    MIN_COL     = 3.0

    merged_set: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        for row_cells in rng.cells:
            cell = row_cells if isinstance(row_cells, tuple) else (row_cells.row, row_cells.column)
            merged_set.add(cell)

    content_w: dict[int, float] = {}
    for ci in range(1, num_cols + 1):
        max_len = MIN_COL
        for ri in range(1, num_data_rows + 2):
            if (ri, ci) in merged_set:
                continue
            val = ws.cell(row=ri, column=ci).value
            if val is None:
                continue
            text   = str(val)
            latin  = sum(1 for ch in text if ord(ch) < 0x0600)
            arabic = len(text) - latin
            max_len = max(max_len, latin + arabic * 1.4)
        content_w[ci] = max_len

    dispatch_w = max(MIN_COL, content_w.get(dispatch_col, MIN_COL))
    remaining  = TOTAL_WIDTH - dispatch_w
    other_cols = [ci for ci in range(1, num_cols + 1) if ci != dispatch_col]
    other_total = sum(max(content_w[ci], MIN_COL) for ci in other_cols)
    scale       = remaining / other_total if other_total > 0 else 1.0

    for ci in range(1, num_cols + 1):
        w = dispatch_w if ci == dispatch_col else max(MIN_COL, content_w[ci]) * scale
        ws.column_dimensions[get_column_letter(ci)].width = round(w, 2)


# ── Sheet writers ─────────────────────────────────────────────────────────────

def _write_title_rows(ws: Worksheet, today_str: str, sector: str) -> None:
    """Rows 1-3: institution title, report title+sector+date, column headers."""
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value=CENTER_TITLE)
    c.font = Font(bold=True, name="Arial", size=14, color=C_BLACK)
    c.fill = WHITE_FILL; c.alignment = _al("center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:G2")
    c = ws.cell(row=2, column=1,
                value=f"بيان تسليم الارساليات الصادرة  ◈  {sector}  ◈  {today_str}")
    c.font = Font(bold=True, name="Arial", size=12, color=C_BLACK)
    c.fill = ODD_FILL; c.alignment = _al("center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)
    ch = ws.cell(row=3, column=1, value=NOTES_MERGE_HEADER)
    ch.font = HDR_FONT; ch.fill = HEADER_FILL; ch.alignment = _al("center")

    for ci, header in enumerate(["الوزن", "رقم الارسالية", "اسم مركز الحركة", "الكود", "م"], 3):
        c = ws.cell(row=3, column=ci, value=header)
        c.font = HDR_FONT; c.fill = HEADER_FILL; c.alignment = _al("center")
    ws.row_dimensions[3].height = 20


def _write_data_rows(ws: Worksheet, sdf: pd.DataFrame) -> None:
    """Write one data row per shipment starting at row 4."""
    for i, row in sdf.iterrows():
        er        = i + 4 # type: ignore
        fill      = ALT_FILL if i % 2 == 0 else ODD_FILL # type: ignore
        is_single = row["total_items"] == 1

        c1 = ws.cell(row=er, column=1, value="على المكشوف" if is_single else "")
        c1.font   = DATA_FONT
        c1.fill   = fill if is_single else EMPTY_CELL_FILL
        c1.border = BORDER; c1.alignment = _al("center")

        c2 = ws.cell(row=er, column=2, value="قابل للكسر")
        c2.font   = DATA_FONT; c2.fill = fill
        c2.border = BORDER;    c2.alignment = _al("center")

        for ci, val in enumerate(
            [row["weight"], row["dispatch_no"],
             row["office_name"], row["office_code"], i + 1], 3 # type: ignore
        ):
            c = ws.cell(row=er, column=ci, value=val)
            c.font = DATA_FONT; c.fill = fill
            c.border = BORDER;  c.alignment = _al("center")

        ws.row_dimensions[er].height = 18


def _write_summary_row(ws: Worksheet, sr: int,
                       total_weight: float, total_count: int,
                       num_offices: int) -> None:
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=2)
    _bold = Font(bold=True, name="Arial", size=12, color=C_BLACK)

    c1 = ws.cell(row=sr, column=1,
                 value=f"إجمالي الوزن\n{total_weight:.3f} كجم")
    c1.font = _bold; c1.fill = ODD_FILL
    c1.border = THICK_BORDER; c1.alignment = _al("center", wrap=True)

    ws.merge_cells(start_row=sr, start_column=3, end_row=sr, end_column=4)
    c2 = ws.cell(row=sr, column=3, value=f"عدد المراكز\n{num_offices}")
    c2.font = _bold; c2.fill = ODD_FILL
    c2.border = THICK_BORDER; c2.alignment = _al("center", wrap=True)

    ws.merge_cells(start_row=sr, start_column=5, end_row=sr, end_column=7)
    c3 = ws.cell(row=sr, column=5, value=f"إجمالي الارساليات\n{total_count}")
    c3.font = _bold; c3.fill = ODD_FILL
    c3.border = THICK_BORDER; c3.alignment = _al("center", wrap=True)
    ws.row_dimensions[sr].height = 38


def _write_signature_row(ws: Worksheet, sig: int,
                         end_col: int, height: int = 28) -> None:
    ws.merge_cells(start_row=sig, start_column=1,
                   end_row=sig, end_column=end_col)
    sb = ws.cell(row=sig, column=1,
                 value="توقيع المستلم :  .................................")
    sb.font      = Font(bold=True, name="Arial", size=14, color=C_BLACK)
    sb.fill      = ODD_FILL
    sb.alignment = _al("right", wrap=False)
    ws.row_dimensions[sig].height = height


def _set_sector_col_widths(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 7


def write_sector(wb: Workbook, sdf: pd.DataFrame,
                 sector: str) -> tuple[int, float]:
    """
    Write one sector sheet and return (total_count, total_weight).
    """
    ws = wb.create_sheet(title=sector)
    ws.sheet_view.rightToLeft = True
    today_str = date.today().strftime("%Y-%m-%d")

    _write_title_rows(ws, today_str, sector)

    sdf = (
        sdf.copy()
        .assign(_sort_no=lambda d: pd.to_numeric(d["dispatch_no"], errors="coerce"))
        .sort_values(["office_name", "_sort_no"], ascending=[True, True])
        .reset_index(drop=True)
    )
    _write_data_rows(ws, sdf)

    total_count  = len(sdf)
    total_weight = sdf["weight"].sum()
    num_offices  = sdf["office_name"].nunique()
    sr           = total_count + 4

    _write_summary_row(ws, sr, total_weight, total_count, num_offices)
    _write_signature_row(ws, sig=sr + 1, end_col=7)

    ws.freeze_panes = "A4"
    _auto_col_widths(ws, num_cols=7, num_data_rows=sr + 1)
    _set_sector_col_widths(ws)
    _print_setup(ws, row_header_end=3)
    _protect(ws)

    return total_count, total_weight


def write_ijmaly(wb: Workbook, sector_totals: dict) -> None:
    """Write the الاجمالى summary sheet."""
    ws = wb.create_sheet(title="الاجمالى")
    ws.sheet_view.rightToLeft = True
    today_str = date.today().strftime("%Y-%m-%d")

    # Title rows
    for row, value, fill in [
        (1, CENTER_TITLE,                       WHITE_FILL),
        (2, "بيان اجمالى الارساليات الصادرة",  ODD_FILL),
        (3, f"التاريخ :  {today_str}",          WHITE_FILL),
    ]:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=value)
        c.font = DATE_FONT; c.fill = fill; c.alignment = _al("center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16

    # Column headers
    for ci, header in enumerate(["الوزن (كجم)", "عدد الارساليات", "القطاع", "م"], 1):
        c = ws.cell(row=4, column=ci, value=header)
        c.font = HDR_FONT; c.fill = HEADER_FILL; c.alignment = _al("center")
    ws.row_dimensions[4].height = 18

    # Data rows
    grand_count = 0; grand_weight = 0.0
    for i, sector in enumerate(SECTOR_SHEETS, 1):
        info = sector_totals.get(sector, {"count": 0, "weight": 0.0})
        er   = i + 4
        fill = ALT_FILL if i % 2 == 0 else ODD_FILL
        grand_count  += info["count"]
        grand_weight += info["weight"]
        for ci, val in enumerate(
            [round(info["weight"], 3), info["count"], sector, i], 1
        ):
            c = ws.cell(row=er, column=ci, value=val)
            c.font = DATA_FONT; c.fill = fill
            c.border = BORDER;  c.alignment = _al("center")
        ws.row_dimensions[er].height = 18

    # Totals row
    tr = len(SECTOR_SHEETS) + 5
    ws.merge_cells(start_row=tr, start_column=3, end_row=tr, end_column=4)
    for ci, val in enumerate([round(grand_weight, 3), grand_count, "الاجمالى"], 1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = SUM_FONT; c.fill = HEADER_FILL
        c.border = THICK_BORDER; c.alignment = _al("center")
    ws.row_dimensions[tr].height = 22

    _write_signature_row(ws, sig=tr + 1, end_col=4, height=30)

    sig = tr + 1
    _auto_col_widths(ws, num_cols=4, num_data_rows=sig)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 7

    _print_setup(ws, row_header_end=4)
    _protect(ws)


# ── Public entry point ────────────────────────────────────────────────────────

def build_workbook(
    raw_df: pd.DataFrame,
    filtered_df: pd.DataFrame,
    output_dir: str,
    log_fn,
    after_6pm: bool = False,
) -> tuple[str, dict]:
    """
    Build the full output workbook and save it.

    Returns:
        out_path       — absolute path of the saved file
        sector_totals  — { sector: {"count": int, "weight": float} }
    """
    log_fn("📝 إنشاء ملف Excel...")
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore

    sector_totals: dict = {}
    for sector in SECTOR_SHEETS:
        sdf = filtered_df[filtered_df["sector"] == sector].copy()
        log_fn(f"   📊 {sector}: {len(sdf)} ارسالية")
        cnt, wgt = write_sector(wb, sdf, sector)
        sector_totals[sector] = {"count": cnt, "weight": wgt}

    write_ijmaly(wb, sector_totals)

    save_dir = Path(output_dir)
    save_dir.mkdir(parents=True, exist_ok=True)
    today  = date.today().strftime("%Y-%m-%d")
    suffix = " - بعد 6 مساءً" if after_6pm else ""
    out_path = save_dir / f"بيان تسليم الارساليات الصادرة - {today}{suffix}.xlsx"

    if out_path.exists():
        try:
            out_path.unlink()
        except OSError:
            raise RuntimeError(
                f"الملف مفتوح في Excel، يرجى إغلاقه أولاً:\n{out_path}"
            )

    wb.save(out_path)
    log_fn(f"✅ تم الحفظ:\n{out_path}")
    return str(out_path), sector_totals