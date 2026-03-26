"""
config/styles.py — All openpyxl fonts, fills, and borders.
Optimized for BLACK & WHITE printing — all grays print identically in B&W and color.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Colors ────────────────────────────────────────────────────────────────────
C_BLACK      = "000000"
C_WHITE      = "FFFFFF"
C_DARK       = "1A1A1A"   # near-black  — main headers bg
C_MID_DARK   = "404040"   # dark gray   — sub-header bg
C_MID        = "787878"   # mid gray    — summary bg
C_LIGHT_GRAY = "D0D0D0"   # light gray  — alt rows
C_VERY_LIGHT = "F2F2F2"   # near-white  — odd rows / sig bg
C_EMPTY_CELL = "BBBBBB"   # medium gray — empty note cell

# Aliases kept for backward compatibility
C_NAVY = C_DARK

# ── Fills ─────────────────────────────────────────────────────────────────────
def _fill(color: str) -> PatternFill:
    return PatternFill("solid", start_color=color, end_color=color)

HEADER_FILL     = _fill(C_DARK)
SUBHDR_FILL     = _fill(C_MID_DARK)
ALT_FILL        = _fill(C_LIGHT_GRAY)
WHITE_FILL      = _fill(C_WHITE)
ODD_FILL        = _fill(C_VERY_LIGHT)
SUM_FILL        = _fill(C_MID)
SIG_FILL        = _fill(C_VERY_LIGHT)
EMPTY_CELL_FILL = _fill(C_EMPTY_CELL)
STAT_FILL       = _fill(C_DARK)

# Aliases
NAVY_FILL    = HEADER_FILL
GOLD_BG_FILL = SUBHDR_FILL

# ── Fonts (all black — print-safe) ────────────────────────────────────────────
BIG_FONT    = Font(bold=True,   color=C_WHITE, name="Zain", size=14)
HDR_FONT    = Font(bold=True,   color=C_WHITE, name="Zain", size=12)
SUBHDR_FONT = Font(bold=True,   color=C_WHITE, name="Zain",            size=12)
GOLD_FONT   = Font(bold=True,   color=C_WHITE, name="Zain",            size=12)
SUM_FONT    = Font(bold=True,   color=C_WHITE, name="Zain",            size=12)
DATA_FONT   = Font(             color=C_BLACK, name="Tajawal",          size=12)
SIG_FONT    = Font(italic=True, color=C_DARK,  name="Zain",            size=9)
DATE_FONT   = Font(bold=True,   color=C_BLACK, name="Zain", size=14)
STAT_FONT   = Font(bold=True,   color=C_WHITE, name="Zain",            size=12)

# ── Borders (all black — print-safe) ─────────────────────────────────────────
_THIN  = Side(style="thin",   color=C_BLACK)
_THICK = Side(style="medium", color=C_BLACK)

BORDER       = Border(left=_THIN,  right=_THIN,  top=_THIN,  bottom=_THIN)
THICK_BORDER = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THICK)
BOT_THICK    = Border(left=_THIN,  right=_THIN,  top=_THIN,  bottom=_THICK)
TOP_THICK    = Border(left=_THICK, right=_THICK, top=_THICK, bottom=_THIN)

# Keep old names for any code that imported THIN / THICK directly
THIN  = _THIN
THICK = _THICK