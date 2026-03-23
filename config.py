"""
config.py — All constants, column indices, colors, and styles.
Optimized for BLACK & WHITE printing.
"""

from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────
# SOURCE FILE — COLUMN INDICES (0-based, header=1)
# ─────────────────────────────────────────────
COL_DEST_OFFICE = 4
COL_DISPATCH_NO = 6
COL_DATETIME    = 8
COL_STATUS      = 9
COL_TOTAL_ITEMS = 12
COL_WEIGHT      = 14

VALID_STATUSES = {"closed"}   # only Closed rows are processed

# ─────────────────────────────────────────────
# REPORT SETTINGS
# ─────────────────────────────────────────────
CENTER_TITLE       = "المركز اللوجيستى بالمنصورة 9900"
SECTOR_SHEETS      = ["بحرى", "قبلى", "رمسيس", "طنطا", "المنصورة"]
OUT_HEADERS        = ["م", "الكود", "اسم مركز الحركة", "رقم الارسالية",
                      "الوزن", "قابل للكسر", "على المكشوف"]
NOTES_MERGE_HEADER = "ملاحظات"

# ─────────────────────────────────────────────
# DEFAULT SAVE PATH
# ─────────────────────────────────────────────
DEFAULT_SAVE_DIR = Path.home() / "Desktop" / "التقارير"

# ─────────────────────────────────────────────
# B&W PRINT-SAFE COLOR PALETTE
# All grays — print identically in B&W and color
# ─────────────────────────────────────────────
C_BLACK      = "000000"
C_WHITE      = "FFFFFF"
C_DARK       = "1A1A1A"   # near-black  — main headers bg
C_MID_DARK   = "404040"   # dark gray   — sub-header bg
C_MID        = "787878"   # mid gray    — summary bg
C_LIGHT_GRAY = "D0D0D0"   # light gray  — alt rows
C_VERY_LIGHT = "F2F2F2"   # near-white  — odd rows / sig bg
C_EMPTY_CELL = "BBBBBB"   # medium gray — empty note cell

# ─────────────────────────────────────────────
# FILLS
# ─────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", start_color=C_DARK,       end_color=C_DARK)
SUBHDR_FILL    = PatternFill("solid", start_color=C_MID_DARK,   end_color=C_MID_DARK)
ALT_FILL       = PatternFill("solid", start_color=C_LIGHT_GRAY, end_color=C_LIGHT_GRAY)
WHITE_FILL     = PatternFill("solid", start_color=C_WHITE,      end_color=C_WHITE)
ODD_FILL       = PatternFill("solid", start_color=C_VERY_LIGHT, end_color=C_VERY_LIGHT)
SUM_FILL       = PatternFill("solid", start_color=C_MID,        end_color=C_MID)
SIG_FILL       = PatternFill("solid", start_color=C_VERY_LIGHT, end_color=C_VERY_LIGHT)
EMPTY_CELL_FILL= PatternFill("solid", start_color=C_EMPTY_CELL, end_color=C_EMPTY_CELL)
STAT_FILL      = PatternFill("solid", start_color=C_DARK,       end_color=C_DARK)

# Aliases kept for backward compatibility
NAVY_FILL    = HEADER_FILL
GOLD_BG_FILL = SUBHDR_FILL

# ─────────────────────────────────────────────
# FONTS  (all black — print-safe)
# ─────────────────────────────────────────────
BIG_FONT   = Font(bold=True, color=C_WHITE,      name="Arial", size=14)
HDR_FONT   = Font(bold=True, color=C_WHITE,      name="Arial", size=12)
SUBHDR_FONT= Font(bold=True, color=C_WHITE,      name="Arial", size=12)
GOLD_FONT  = Font(bold=True, color=C_WHITE,      name="Arial", size=12)
SUM_FONT   = Font(bold=True, color=C_WHITE,      name="Arial", size=12)
DATA_FONT  = Font(           color=C_BLACK,      name="Arial", size=11)
SIG_FONT   = Font(italic=True,color=C_DARK,      name="Arial", size=9)
DATE_FONT  = Font(bold=True, color=C_BLACK,      name="Arial", size=12)
STAT_FONT  = Font(bold=True, color=C_WHITE,      name="Arial", size=12)

# ─────────────────────────────────────────────
# BORDERS  (all black — print-safe)
# ─────────────────────────────────────────────
THIN         = Side(style="thin",   color=C_BLACK)
THICK        = Side(style="medium", color=C_BLACK)
BORDER       = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
THICK_BORDER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
BOT_THICK    = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THICK)
TOP_THICK    = Border(left=THICK, right=THICK, top=THICK, bottom=THIN)

# C_NAVY alias used in excel_writer
C_NAVY = C_DARK