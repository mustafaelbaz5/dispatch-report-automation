"""
gui/dialogs/print_dialog.py — Dialog for selecting and printing sector sheets.
"""

import os
import sys
import subprocess
import tempfile
from pathlib import Path
from tkinter import messagebox

import customtkinter as ctk

from config import SECTOR_SHEETS
from gui.theme import (
    BG_CARD, BG_INPUT, BG_HOVER, BORDER, BORDER_LT,
    G, G_H, TEXT_PRI, TEXT_SEC,
    font,
)
from gui.widgets import Divider, GhostBtn, PrimaryBtn


class PrintDialog(ctk.CTkToplevel):
    """Let the user pick which sector sheets to send to the printer."""

    def __init__(self, parent, output_path: str):
        super().__init__(parent)
        self.title("طباعة الشيتات")
        self.geometry("420x480")
        self.resizable(False, False)
        self.configure(fg_color=BG_CARD)
        self.lift(); self.focus_force(); self.grab_set()

        self._path   = output_path
        self._checks: dict[str, ctk.BooleanVar] = {}
        self._build()

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build(self) -> None:
        ctk.CTkFrame(self, fg_color=G, height=3, corner_radius=0).pack(fill="x")
        self._build_header()
        Divider(self).pack(fill="x", padx=20, pady=(8, 0))
        self._build_select_row()
        self._build_sheet_list()
        Divider(self).pack(fill="x", padx=20, pady=(0, 12))
        self._build_footer()

    def _build_header(self) -> None:
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(16, 8))
        ctk.CTkLabel(hdr, text="🖨  اختر الشيتات للطباعة",
                     font=font(15, bold=True), text_color=TEXT_PRI,
                     anchor="e").pack(anchor="e")
        ctk.CTkLabel(hdr, text=Path(self._path).name,
                     font=font(9), text_color=TEXT_SEC,
                     anchor="e", wraplength=380).pack(anchor="e", pady=(2, 0))

    def _build_select_row(self) -> None:
        row = ctk.CTkFrame(self, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(8, 6))
        for label, state in [("تحديد الكل", True), ("إلغاء الكل", False)]:
            ctk.CTkButton(row, text=label, font=font(9),
                          fg_color=BG_INPUT, hover_color=BG_HOVER,
                          text_color=TEXT_SEC, border_width=1,
                          border_color=BORDER, height=26, width=84,
                          corner_radius=6,
                          command=lambda s=state: self._select_all(s)
                          ).pack(side="left", padx=(0, 6))

    def _build_sheet_list(self) -> None:
        sf = ctk.CTkScrollableFrame(self, fg_color=BG_INPUT,
                                    corner_radius=10, border_width=1,
                                    border_color=BORDER, height=200)
        sf.pack(fill="x", padx=20, pady=(0, 12))

        for sheet in SECTOR_SHEETS + ["الاجمالى"]:
            var = ctk.BooleanVar(value=True)
            self._checks[sheet] = var
            row = ctk.CTkFrame(sf, fg_color="transparent")
            row.pack(fill="x", padx=6, pady=3)
            ctk.CTkCheckBox(row, text=sheet, variable=var,
                            font=font(12), text_color=TEXT_PRI,
                            checkmark_color="#fff", fg_color=G,
                            hover_color=G_H, border_color=BORDER_LT,
                            corner_radius=5).pack(side="right", padx=4)

    def _build_footer(self) -> None:
        row = ctk.CTkFrame(self, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(0, 18))
        GhostBtn(row, text="إلغاء", cmd=self.destroy,
                 accent=TEXT_SEC, height=42, icon="✕"
                 ).pack(side="left", padx=(0, 8))
        PrimaryBtn(row, text="طباعة", cmd=self._print,
                   icon="🖨", height=42
                   ).pack(side="right", fill="x", expand=True)

    # ── Actions ───────────────────────────────────────────────────────────────

    def _select_all(self, state: bool) -> None:
        for var in self._checks.values():
            var.set(state)

    def _print(self) -> None:
        selected = [s for s, v in self._checks.items() if v.get()]
        if not selected:
            messagebox.showwarning("تنبيه", "يرجى تحديد شيت واحد على الأقل.", parent=self)
            return
        if not Path(self._path).exists():
            messagebox.showerror("خطأ", "ملف الإخراج غير موجود.", parent=self)
            return
        try:
            _print_selected_sheets(self._path, selected)
            self.destroy()
            messagebox.showinfo("الطباعة", f"✅ تم إرسال {len(selected)} شيت للطابعة.")
        except Exception as e:
            messagebox.showerror("خطأ في الطباعة", str(e), parent=self)


# ── Printing logic (no GUI dependency) ───────────────────────────────────────

def _print_selected_sheets(source_path: str, sheets: list[str]) -> None:
    import openpyxl
    wb = openpyxl.load_workbook(source_path)
    for name in list(wb.sheetnames):
        if name not in sheets:
            del wb[name]
    if not wb.sheetnames:
        raise RuntimeError("لم يتم العثور على الشيتات المحددة.")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="print_")
    tmp.close()
    wb.save(tmp.name)

    if sys.platform == "win32":
        os.startfile(tmp.name)
    else:
        subprocess.run(["lp", tmp.name], check=True)