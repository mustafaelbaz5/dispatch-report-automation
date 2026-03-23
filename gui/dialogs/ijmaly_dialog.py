"""
gui/dialogs/ijmaly_dialog.py — Dialog for printing the الاجمالى summary sheet.
"""

import os
import sys
import subprocess
import tempfile
from datetime import date
from tkinter import messagebox

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import SECTOR_SHEETS, CENTER_TITLE
from config.styles import (
    HDR_FONT, SUM_FONT, DATA_FONT, DATE_FONT,
    HEADER_FILL, ALT_FILL, ODD_FILL,
    BORDER, THICK_BORDER,
    C_BLACK,
)
from gui.theme import (
    BG_CARD, BG_INPUT, BG_HOVER, BORDER as GUI_BORDER, BORDER_LT,
    A, A_H, T, T_DIM,
    TEXT_PRI, TEXT_SEC, TEXT_GOLD,
    font,
)
from gui.widgets import Divider, GhostBtn, PrimaryBtn


class IjmalyPrintDialog(ctk.CTkToplevel):
    """Let the user pick which sectors to include in the الاجمالى printout."""

    def __init__(self, parent, output_path: str, sector_totals: dict):
        super().__init__(parent)
        self.title("طباعة الاجمالى")
        self.geometry("440x500")
        self.resizable(False, False)
        self.configure(fg_color=BG_CARD)
        self.lift(); self.focus_force(); self.grab_set()

        self._path          = output_path
        self._sector_totals = sector_totals
        self._checks: dict[str, ctk.BooleanVar] = {}
        self._build()

    # ── Build ─────────────────────────────────────────────────────────────────

    def _build(self) -> None:
        ctk.CTkFrame(self, fg_color=A, height=3, corner_radius=0).pack(fill="x")
        self._build_header()
        Divider(self).pack(fill="x", padx=20, pady=(8, 0))
        self._build_select_row()
        self._build_sector_list()
        Divider(self).pack(fill="x", padx=20, pady=(0, 12))
        self._build_footer()

    def _build_header(self) -> None:
        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=20, pady=(16, 8))
        ctk.CTkLabel(hdr, text="📊  اختر قطاعات الاجمالى",
                     font=font(15, bold=True), text_color=TEXT_PRI,
                     anchor="e").pack(anchor="e")
        ctk.CTkLabel(hdr, text="اختر القطاعات التي تريد إظهارها في جدول الاجمالى",
                     font=font(9), text_color=TEXT_SEC,
                     anchor="e").pack(anchor="e")

    def _build_select_row(self) -> None:
        row = ctk.CTkFrame(self, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(8, 6))
        for label, state in [("تحديد الكل", True), ("إلغاء الكل", False)]:
            ctk.CTkButton(row, text=label, font=font(9),
                          fg_color=BG_INPUT, hover_color=BG_HOVER,
                          text_color=TEXT_SEC, border_width=1,
                          border_color=GUI_BORDER, height=26, width=84,
                          corner_radius=6,
                          command=lambda s=state: self._select_all(s)
                          ).pack(side="left", padx=(0, 6))

    def _build_sector_list(self) -> None:
        sf = ctk.CTkScrollableFrame(self, fg_color=BG_INPUT,
                                    corner_radius=10, border_width=1,
                                    border_color=GUI_BORDER, height=230)
        sf.pack(fill="x", padx=20, pady=(0, 12))

        for sector in SECTOR_SHEETS:
            info = self._sector_totals.get(sector, {"count": 0, "weight": 0.0})
            var  = ctk.BooleanVar(value=True)
            self._checks[sector] = var

            row = ctk.CTkFrame(sf, fg_color="transparent")
            row.pack(fill="x", padx=6, pady=4)

            pill = ctk.CTkFrame(row, fg_color=T_DIM, corner_radius=6)
            pill.pack(side="left")
            ctk.CTkLabel(pill,
                         text=f"{info['count']} ·  {info['weight']:.1f} كجم",
                         font=font(9), text_color=T).pack(padx=8, pady=4)

            ctk.CTkCheckBox(row, text=sector, variable=var,
                            font=font(12, bold=True), text_color=TEXT_PRI,
                            checkmark_color="#fff", fg_color=A,
                            hover_color=A_H, border_color=BORDER_LT,
                            corner_radius=5).pack(side="right", padx=4)

    def _build_footer(self) -> None:
        row = ctk.CTkFrame(self, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(0, 18))
        GhostBtn(row, text="إلغاء", cmd=self.destroy,
                 accent=TEXT_SEC, height=42, icon="✕"
                 ).pack(side="left", padx=(0, 8))
        PrimaryBtn(row, text="طباعة", cmd=self._print,
                   color=A, hcolor=A_H, icon="🖨", height=42
                   ).pack(side="right", fill="x", expand=True)

    # ── Actions ───────────────────────────────────────────────────────────────

    def _select_all(self, state: bool) -> None:
        for var in self._checks.values():
            var.set(state)

    def _print(self) -> None:
        selected = [s for s, v in self._checks.items() if v.get()]
        if not selected:
            messagebox.showwarning("تنبيه", "يرجى تحديد قطاع واحد على الأقل.", parent=self)
            return
        try:
            _build_and_print_ijmaly(selected, self._sector_totals)
            self.destroy()
            messagebox.showinfo("الطباعة", "✅ تم إرسال الاجمالى للطابعة.")
        except Exception as e:
            messagebox.showerror("خطأ في الطباعة", str(e), parent=self)


# ── Printing logic (no GUI dependency) ───────────────────────────────────────

def _al(h: str = "center", v: str = "center",
        ro: int = 2, wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, readingOrder=ro, wrap_text=wrap)


def _build_and_print_ijmaly(selected_sectors: list[str],
                             sector_totals: dict) -> None:
    today_str = date.today().strftime("%Y-%m-%d")
    wb = Workbook()
    ws = wb.active
    ws.title = "الاجمالى"
    ws.sheet_view.rightToLeft = True

    # Title rows
    for row, value, fill in [
        (1, CENTER_TITLE,                       PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")),
        (2, "بيان اجمالى الارساليات الصادرة",  PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")),
        (3, f"التاريخ :  {today_str}",          PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")),
    ]:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row=row, column=1, value=value)
        c.font = Font(bold=True, name="Arial", size=13 if row == 2 else 14, color=C_BLACK)
        c.fill = fill; c.alignment = _al("center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16

    # Column headers
    for ci, header in enumerate(["الوزن (كجم)", "عدد الارساليات", "القطاع", "م"], 1):
        c = ws.cell(row=4, column=ci, value=header)
        c.font = HDR_FONT; c.fill = HEADER_FILL; c.alignment = _al("center")
    ws.row_dimensions[4].height = 18

    # Data rows
    grand_count = 0; grand_weight = 0.0
    for i, sector in enumerate(selected_sectors, 1):
        info = sector_totals.get(sector, {"count": 0, "weight": 0.0})
        er   = i + 4
        fill = ALT_FILL if i % 2 == 0 else ODD_FILL
        grand_count  += info["count"]
        grand_weight += info["weight"]
        for ci, val in enumerate([round(info["weight"], 3), info["count"], sector, i], 1):
            c = ws.cell(row=er, column=ci, value=val)
            c.font = DATA_FONT; c.fill = fill
            c.border = BORDER; c.alignment = _al("center")
        ws.row_dimensions[er].height = 18

    # Totals row
    tr = len(selected_sectors) + 5
    ws.merge_cells(start_row=tr, start_column=3, end_row=tr, end_column=4)
    for ci, val in enumerate([round(grand_weight, 3), grand_count, "الاجمالى"], 1):
        c = ws.cell(row=tr, column=ci, value=val)
        c.font = SUM_FONT; c.fill = HEADER_FILL
        c.border = THICK_BORDER; c.alignment = _al("center")
    ws.row_dimensions[tr].height = 22

    # Signature row
    sig = tr + 1
    ws.merge_cells(start_row=sig, start_column=1, end_row=sig, end_column=4)
    sb = ws.cell(row=sig, column=1, value="توقيع المستلم :  .................................")
    sb.font      = Font(bold=True, name="Arial", size=14, color=C_BLACK)
    sb.fill      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    sb.alignment = _al("right")
    ws.row_dimensions[sig].height = 30

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 7
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left = ws.page_margins.right  = 0.30
    ws.page_margins.top  = ws.page_margins.bottom = 0.30
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:4"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="ijmaly_print_")
    tmp.close(); wb.save(tmp.name)

    if sys.platform == "win32":
        os.startfile(tmp.name)
    else:
        subprocess.run(["lp", tmp.name], check=True)