"""
app.py — Main GUI entry point.
Run: python app.py

Requirements:
    pip install customtkinter pandas openpyxl tkinterdnd2
"""

import os
import threading
import subprocess
import sys
from pathlib import Path

import customtkinter as ctk
from tkinter import filedialog, messagebox

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    _DND_AVAILABLE = True
except ImportError:
    _DND_AVAILABLE = False

from config import DEFAULT_SAVE_DIR, SECTOR_SHEETS
from processor import load_and_filter
from excel_writer import build_workbook


# ─────────────────────────────────────────────
# APPEARANCE
# ─────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

G       = "#10B981"
G_H     = "#059669"
G_DIM   = "#064E3B"

T       = "#38BDF8"
T_H     = "#0EA5E9"
T_DIM   = "#0C4A6E"

A       = "#818CF8"
A_H     = "#6366F1"

BG_APP    = "#080B10"
BG_CARD   = "#0F1520"
BG_INPUT  = "#141C28"
BG_HDR    = "#090D16"
BG_HOVER  = "#1A2535"

BORDER_C  = "#1E2D42"
BORDER_L  = "#2A3F58"

TEXT_PRI  = "#E2E8F0"
TEXT_SEC  = "#7A90A8"
TEXT_DIM  = "#384D62"
TEXT_GOLD = "#FCD34D"

RED       = "#F87171"
AMBER     = "#FBBF24"
SUCCESS   = "#34D399"


def _font(size=12, bold=False, family="Tajawal"):
    return ctk.CTkFont(family=family, size=size,
                       weight="bold" if bold else "normal")


def _mono(size=11):
    return ctk.CTkFont(family="Consolas", size=size)


# ─────────────────────────────────────────────
# REUSABLE WIDGETS
# ─────────────────────────────────────────────
class SectionLabel(ctk.CTkLabel):
    def __init__(self, parent, text, size=11, color=TEXT_SEC, **kw):
        super().__init__(parent, text=text, font=_font(size),
                         text_color=color, anchor="e", justify="right", **kw)


class Divider(ctk.CTkFrame):
    def __init__(self, parent, **kw):
        super().__init__(parent, fg_color=BORDER_C, height=1, **kw)


class GlowButton(ctk.CTkButton):
    def __init__(self, parent, text, cmd,
                 color=G, hcolor=None, text_color="#FFFFFF",
                 height=48, icon="", **kw):
        label = f"{icon}  {text}" if icon else text
        super().__init__(
            parent, text=label, font=_font(13, bold=True),
            fg_color=color, hover_color=hcolor or G_H,
            text_color=text_color, height=height, corner_radius=12,
            border_width=1, border_color=color, command=cmd, **kw)


class OutlineButton(ctk.CTkButton):
    def __init__(self, parent, text, cmd,
                 accent=T, height=40, icon="", **kw):
        label = f"{icon}  {text}" if icon else text
        super().__init__(
            parent, text=label, font=_font(12, bold=True),
            fg_color="transparent", hover_color=BG_HOVER,
            text_color=accent, height=height, corner_radius=10,
            border_width=1, border_color=accent, command=cmd, **kw)


# ─────────────────────────────────────────────
# SECTOR SHEETS PRINT DIALOG
# ─────────────────────────────────────────────
class PrintDialog(ctk.CTkToplevel):
    def __init__(self, parent, output_path: str):
        super().__init__(parent)
        self.title("اختر الشيتات للطباعة")
        self.geometry("460x520")
        self.resizable(False, False)
        self.configure(fg_color=BG_CARD)
        self.grab_set(); self.lift(); self.focus_force()
        self._path   = output_path
        self._checks = {}
        self._build()

    def _build(self):
        ctk.CTkFrame(self, fg_color=G, height=3, corner_radius=0).pack(fill="x")

        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=24, pady=(20, 4))
        ctk.CTkLabel(hdr, text="🖨  اختر الشيتات للطباعة",
                     font=_font(17, bold=True), text_color=TEXT_PRI,
                     anchor="e").pack(anchor="e")
        ctk.CTkLabel(hdr, text=Path(self._path).name,
                     font=_font(10), text_color=TEXT_DIM,
                     anchor="e", wraplength=400).pack(anchor="e", pady=(2, 0))

        Divider(self).pack(fill="x", padx=24, pady=(12, 0))

        sel_row = ctk.CTkFrame(self, fg_color="transparent")
        sel_row.pack(fill="x", padx=24, pady=(10, 6))
        for lbl, state in [("تحديد الكل", True), ("إلغاء الكل", False)]:
            ctk.CTkButton(sel_row, text=lbl, font=_font(10),
                          fg_color=BG_INPUT, hover_color=BG_HOVER,
                          text_color=TEXT_SEC, border_width=1,
                          border_color=BORDER_C, height=28, width=90,
                          corner_radius=8,
                          command=lambda s=state: self._select_all(s)
                          ).pack(side="left", padx=(0, 6))

        sf = ctk.CTkScrollableFrame(self, fg_color=BG_INPUT, corner_radius=12,
                                    border_width=1, border_color=BORDER_C,
                                    height=220)
        sf.pack(fill="x", padx=24, pady=(0, 12))

        for sheet in SECTOR_SHEETS + ["الاجمالى"]:
            var = ctk.BooleanVar(value=True)
            self._checks[sheet] = var
            row = ctk.CTkFrame(sf, fg_color="transparent")
            row.pack(fill="x", padx=8, pady=4)
            badge = ctk.CTkFrame(row, fg_color=G_DIM, corner_radius=6,
                                 width=28, height=28)
            badge.pack(side="right", padx=(0, 8))
            badge.pack_propagate(False)
            ctk.CTkLabel(badge, text="📄", font=_font(11)).pack(expand=True)
            ctk.CTkCheckBox(row, text=sheet, variable=var,
                            font=_font(13), text_color=TEXT_PRI,
                            checkmark_color="#FFFFFF", fg_color=G,
                            hover_color=G_H, border_color=BORDER_L,
                            corner_radius=6).pack(side="right", padx=(0, 6))

        Divider(self).pack(fill="x", padx=24, pady=(0, 14))

        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=24, pady=(0, 20))
        OutlineButton(btn_row, text="إلغاء", cmd=self.destroy,
                      accent=TEXT_SEC, height=44, icon="✕"
                      ).pack(side="left", padx=(0, 8))
        GlowButton(btn_row, text="طباعة", cmd=self._print,
                   color=G, hcolor=G_H, height=44, icon="🖨"
                   ).pack(side="right", fill="x", expand=True)

    def _select_all(self, state: bool):
        for var in self._checks.values():
            var.set(state)

    def _print(self):
        selected = [s for s, v in self._checks.items() if v.get()]
        if not selected:
            messagebox.showwarning("تنبيه", "يرجى تحديد شيت واحد على الأقل.",
                                   parent=self)
            return
        if not Path(self._path).exists():
            messagebox.showerror("خطأ", "ملف الإخراج غير موجود.", parent=self)
            return
        try:
            self._do_print(selected)
            self.destroy()
            messagebox.showinfo("الطباعة",
                                f"✅ تم إرسال {len(selected)} شيت للطابعة.")
        except Exception as e:
            messagebox.showerror("خطأ في الطباعة", str(e), parent=self)

    def _do_print(self, sheets: list):
        import openpyxl, tempfile
        wb = openpyxl.load_workbook(self._path)
        for name in list(wb.sheetnames):
            if name not in sheets:
                del wb[name]
        if not wb.sheetnames:
            raise RuntimeError("لم يتم العثور على الشيتات المحددة في الملف.")
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False,
                                          prefix="print_")
        tmp.close()
        wb.save(tmp.name)
        if sys.platform == "win32":
            os.startfile(tmp.name, "print")   # type: ignore
        else:
            subprocess.run(["lp", tmp.name], check=True)


# ─────────────────────────────────────────────
# الاجمالى PRINT DIALOG
# ─────────────────────────────────────────────
class IjmalyPrintDialog(ctk.CTkToplevel):
    """
    Choose which sector rows to keep in الاجمالى table, then print.
    Shows live count/weight badges per sector.
    """

    def __init__(self, parent, output_path: str, sector_totals: dict):
        super().__init__(parent)
        self.title("طباعة الاجمالى")
        self.geometry("500x600")
        self.resizable(False, False)
        self.configure(fg_color=BG_CARD)
        self.grab_set(); self.lift(); self.focus_force()
        self._path          = output_path
        self._sector_totals = sector_totals
        self._checks        = {}
        self._build()

    def _build(self):
        ctk.CTkFrame(self, fg_color=A, height=3, corner_radius=0).pack(fill="x")

        hdr = ctk.CTkFrame(self, fg_color="transparent")
        hdr.pack(fill="x", padx=24, pady=(20, 4))
        ctk.CTkLabel(hdr, text="📊  طباعة بيان الاجمالى",
                     font=_font(17, bold=True), text_color=TEXT_PRI,
                     anchor="e").pack(anchor="e")
        ctk.CTkLabel(hdr,
                     text="اختر القطاعات التي تريد إظهارها في الجدول",
                     font=_font(10), text_color=TEXT_DIM,
                     anchor="e").pack(anchor="e", pady=(2, 0))

        Divider(self).pack(fill="x", padx=24, pady=(12, 0))

        sel_row = ctk.CTkFrame(self, fg_color="transparent")
        sel_row.pack(fill="x", padx=24, pady=(10, 6))
        for lbl, state in [("تحديد الكل", True), ("إلغاء الكل", False)]:
            ctk.CTkButton(sel_row, text=lbl, font=_font(10),
                          fg_color=BG_INPUT, hover_color=BG_HOVER,
                          text_color=TEXT_SEC, border_width=1,
                          border_color=BORDER_C, height=28, width=90,
                          corner_radius=8,
                          command=lambda s=state: self._select_all(s)
                          ).pack(side="left", padx=(0, 6))

        sf = ctk.CTkScrollableFrame(self, fg_color=BG_INPUT, corner_radius=12,
                                    border_width=1, border_color=BORDER_C,
                                    height=280)
        sf.pack(fill="x", padx=24, pady=(0, 12))

        for sector in SECTOR_SHEETS:
            info = self._sector_totals.get(sector, {"count": 0, "weight": 0.0})
            var  = ctk.BooleanVar(value=True)
            self._checks[sector] = var

            row = ctk.CTkFrame(sf, fg_color="transparent")
            row.pack(fill="x", padx=8, pady=6)

            # stats pill on left side
            pill = ctk.CTkFrame(row, fg_color=T_DIM, corner_radius=8)
            pill.pack(side="left")
            ctk.CTkLabel(pill,
                         text=f"{info['count']} ارسالية  ·  {info['weight']:.1f} كجم",
                         font=_font(9), text_color=T,
                         ).pack(padx=10, pady=5)

            # checkbox on right
            ctk.CTkCheckBox(row, text=sector, variable=var,
                            font=_font(13, bold=True), text_color=TEXT_PRI,
                            checkmark_color="#FFFFFF", fg_color=A,
                            hover_color=A_H, border_color=BORDER_L,
                            corner_radius=6).pack(side="right", padx=(0, 6))

        Divider(self).pack(fill="x", padx=24, pady=(0, 14))

        btn_row = ctk.CTkFrame(self, fg_color="transparent")
        btn_row.pack(fill="x", padx=24, pady=(0, 20))
        OutlineButton(btn_row, text="إلغاء", cmd=self.destroy,
                      accent=TEXT_SEC, height=44, icon="✕"
                      ).pack(side="left", padx=(0, 8))
        GlowButton(btn_row, text="طباعة", cmd=self._print,
                   color=A, hcolor=A_H, height=44, icon="🖨"
                   ).pack(side="right", fill="x", expand=True)

    def _select_all(self, state: bool):
        for var in self._checks.values():
            var.set(state)

    def _print(self):
        selected = [s for s, v in self._checks.items() if v.get()]
        if not selected:
            messagebox.showwarning("تنبيه", "يرجى تحديد قطاع واحد على الأقل.",
                                   parent=self)
            return
        try:
            self._do_print(selected)
            self.destroy()
            messagebox.showinfo("الطباعة", "✅ تم إرسال الاجمالى للطابعة.")
        except Exception as e:
            messagebox.showerror("خطأ في الطباعة", str(e), parent=self)

    def _do_print(self, selected_sectors: list):
        from datetime import date
        import tempfile
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from config import (
            CENTER_TITLE, HDR_FONT, SUM_FONT, DATA_FONT,
            HEADER_FILL, ALT_FILL, ODD_FILL, WHITE_FILL,
            THICK_BORDER, BORDER, C_BLACK,
        )

        def _al(h="center", v="center", ro=2, wrap=False):
            return Alignment(horizontal=h, vertical=v,
                             readingOrder=ro, wrap_text=wrap)

        today_str = date.today().strftime("%Y-%m-%d")
        wb = Workbook()
        ws = wb.active
        ws.title = "الاجمالى"
        ws.sheet_view.rightToLeft = True

        # Row 1 — institution
        ws.merge_cells("A1:D1")
        c = ws.cell(row=1, column=1, value=CENTER_TITLE)
        c.font = Font(bold=True, name="Arial", size=14, color=C_BLACK)
        c.fill = WHITE_FILL; c.alignment = _al("center")
        ws.row_dimensions[1].height = 24

        # Row 2 — title
        ws.merge_cells("A2:D2")
        c = ws.cell(row=2, column=1, value="بيان اجمالى الارساليات الصادرة")
        c.font = Font(bold=True, name="Arial", size=13, color=C_BLACK)
        c.fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
        c.alignment = _al("center")
        ws.row_dimensions[2].height = 20

        # Row 3 — date
        ws.merge_cells("A3:D3")
        c = ws.cell(row=3, column=1, value=f"التاريخ :  {today_str}")
        c.font = Font(bold=True, name="Arial", size=12, color=C_BLACK)
        c.fill = WHITE_FILL; c.alignment = _al("center")
        ws.row_dimensions[3].height = 16

        # Row 4 — column headers
        for ci, h in enumerate(["الوزن (كجم)", "عدد الارساليات", "القطاع", "م"], 1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font = HDR_FONT; c.fill = HEADER_FILL; c.alignment = _al("center")
        ws.row_dimensions[4].height = 18

        # Data rows
        grand_count = 0; grand_weight = 0.0
        for i, sector in enumerate(selected_sectors, 1):
            info = self._sector_totals.get(sector, {"count": 0, "weight": 0.0})
            er   = i + 4
            fill = ALT_FILL if i % 2 == 0 else ODD_FILL
            grand_count  += info["count"]
            grand_weight += info["weight"]
            for ci, val in enumerate(
                [round(info["weight"], 3), info["count"], sector, i], 1
            ):
                c = ws.cell(row=er, column=ci, value=val)
                c.font = DATA_FONT; c.fill = fill
                c.border = BORDER; c.alignment = _al("center")
            ws.row_dimensions[er].height = 18

        # Totals row
        tr = len(selected_sectors) + 5
        ws.merge_cells(start_row=tr, start_column=3, end_row=tr, end_column=4)
        for ci, val in enumerate(
            [round(grand_weight, 3), grand_count, "الاجمالى"], 1
        ):
            c = ws.cell(row=tr, column=ci, value=val)
            c.font = SUM_FONT; c.fill = HEADER_FILL
            c.border = THICK_BORDER; c.alignment = _al("center")
        ws.row_dimensions[tr].height = 22

        # Signature row
        sig = tr + 1
        ws.merge_cells(start_row=sig, start_column=1, end_row=sig, end_column=4)
        sb = ws.cell(row=sig, column=1,
                     value="توقيع المستلم :  .................................")
        sb.font      = Font(bold=True, name="Arial", size=14, color=C_BLACK)
        sb.fill      = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
        sb.alignment = _al("right")
        ws.row_dimensions[sig].height = 30

        # Column widths
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 7

        # Print setup
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize   = 9
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_margins.left = ws.page_margins.right  = 0.30
        ws.page_margins.top  = ws.page_margins.bottom = 0.30
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = "1:4"

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False,
                                          prefix="ijmaly_print_")
        tmp.close()
        wb.save(tmp.name)

        if sys.platform == "win32":
            os.startfile(tmp.name, "print")   # type: ignore
        else:
            subprocess.run(["lp", tmp.name], check=True)


# ─────────────────────────────────────────────
# STATUS BADGE
# ─────────────────────────────────────────────
class StatusBadge(ctk.CTkFrame):
    def __init__(self, parent, **kw):
        super().__init__(parent, fg_color="transparent", **kw)
        self._dot = ctk.CTkFrame(self, width=8, height=8,
                                 corner_radius=4, fg_color=TEXT_DIM)
        self._dot.pack(side="left", padx=(0, 6))
        self._lbl = ctk.CTkLabel(self, text="جاهز",
                                 font=_font(10), text_color=TEXT_DIM)
        self._lbl.pack(side="left")

    def set(self, text: str, color: str):
        self._dot.configure(fg_color=color)
        self._lbl.configure(text=text, text_color=color)


# ─────────────────────────────────────────────
# MAIN APP
# ─────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        if _DND_AVAILABLE:
            try:
                TkinterDnD._require(self)   # type: ignore
            except Exception:
                pass

        self.title("بيان تسليم الارساليات الصادرة")
        self.configure(fg_color=BG_APP)

        # ── Open on left half of screen ──
        self.update_idletasks()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        w  = sw // 2
        h  = sh
        self.geometry(f"{w}x{h}+0+0")
        self.minsize(640, 600)
        self.resizable(True, True)
  

        self._source_path   = ctk.StringVar()
        self._output_dir    = ctk.StringVar(value=str(DEFAULT_SAVE_DIR))
        self._after_6pm     = ctk.BooleanVar(value=False)
        self._last_output   = None
        self._sector_totals = {}

        self._build_ui()

    # ─────────────────────────────────────────
    # UI
    # ─────────────────────────────────────────
    def _build_ui(self):
        self._build_header()
        self._content = ctk.CTkScrollableFrame(
            self, fg_color=BG_APP,
            scrollbar_button_color=BORDER_C,
            scrollbar_button_hover_color=TEXT_SEC)
        self._content.pack(fill="both", expand=True)
        self._build_files_card()
        self._build_options_card()
        self._build_actions_section()
        self._build_log_panel()

    def _build_header(self):
        hdr = ctk.CTkFrame(self, fg_color=BG_HDR, corner_radius=0, height=96)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)

        bar = ctk.CTkFrame(hdr, fg_color="transparent", height=3, corner_radius=0)
        bar.pack(fill="x", side="top")
        bar.pack_propagate(False)
        for color, expand in [(G, True), (T, False), (A, False)]:
            ctk.CTkFrame(bar, fg_color=color, height=3,
                         corner_radius=0).pack(side="left", fill="both",
                                               expand=expand,
                                               ipadx=0 if expand else 40)

        inner = ctk.CTkFrame(hdr, fg_color="transparent")
        inner.place(relx=0.5, rely=0.62, anchor="center")
        ctk.CTkLabel(inner, text="بيان تسليم الارساليات الصادرة",
                     font=_font(20, bold=True), text_color=TEXT_PRI).pack()

        sub = ctk.CTkFrame(inner, fg_color="transparent")
        sub.pack(pady=(3, 0))
        badge = ctk.CTkFrame(sub, fg_color=G_DIM, corner_radius=6,
                             height=20, width=40)
        badge.pack(side="right", padx=(6, 0))
        badge.pack_propagate(False)
        ctk.CTkLabel(badge, text="9900", font=_font(9, bold=True),
                     text_color=SUCCESS).pack(expand=True)
        ctk.CTkLabel(sub, text="المركز اللوجيستى بالمنصورة",
                     font=_font(11), text_color=TEXT_GOLD).pack(side="right")

    def _card(self, parent, title: str, icon: str = "") -> ctk.CTkFrame:
        wrapper = ctk.CTkFrame(parent, fg_color="transparent")
        wrapper.pack(fill="x", padx=20, pady=(14, 0))

        tr = ctk.CTkFrame(wrapper, fg_color="transparent")
        tr.pack(fill="x", pady=(0, 8))
        pills = ctk.CTkFrame(tr, fg_color="transparent")
        pills.pack(side="right", padx=(0, 6))
        for color, h in [(G, 14), (T, 10), (A, 6)]:
            ctk.CTkFrame(pills, fg_color=color, width=3, height=h,
                         corner_radius=2).pack(side="right", padx=1)
        ctk.CTkLabel(tr, text=f"{icon}  {title}" if icon else title,
                     font=_font(13, bold=True), text_color=TEXT_PRI
                     ).pack(side="right")

        card = ctk.CTkFrame(wrapper, fg_color=BG_CARD, corner_radius=14,
                            border_width=1, border_color=BORDER_C)
        card.pack(fill="x")
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=18, pady=16)
        return inner

    def _build_files_card(self):
        inner = self._card(self._content, "الملفات", "📁")

        SectionLabel(inner, "ملف المصدر ( Excel ) :", size=11
                     ).pack(anchor="e", pady=(0, 8))

        self._drop_zone = ctk.CTkFrame(
            inner, fg_color=BG_INPUT, corner_radius=12,
            border_width=2, border_color=BORDER_C, height=160)
        self._drop_zone.pack(fill="x", pady=(0, 10))
        self._drop_zone.pack_propagate(False)

        dz_inner = ctk.CTkFrame(self._drop_zone, fg_color="transparent")
        dz_inner.place(relx=0.5, rely=0.5, anchor="center")
        ctk.CTkLabel(dz_inner, text="⬆", font=_font(22),
                     text_color=TEXT_DIM).pack()
        self._drop_lbl = ctk.CTkLabel(dz_inner,
                                      text="اسحب ملف Excel وأفلته هنا",
                                      font=_font(12, bold=True),
                                      text_color=TEXT_SEC)
        self._drop_lbl.pack(pady=(2, 0))
        self._drop_sub = ctk.CTkLabel(dz_inner, text=".xlsx  ·  .xls  ·  .xlsm",
                                      font=_font(9), text_color=TEXT_DIM)
        self._drop_sub.pack()

        if _DND_AVAILABLE:
            try:
                dz = self._drop_zone._canvas           # type: ignore
                dz.drop_target_register(DND_FILES)     # type: ignore
                dz.dnd_bind("<<Drop>>", self._on_drop) # type: ignore
            except Exception:
                pass
        self._drop_zone.bind("<Enter>", lambda e: self._dz_hover(True))
        self._drop_zone.bind("<Leave>", lambda e: self._dz_hover(False))

        pr = ctk.CTkFrame(inner, fg_color="transparent")
        pr.pack(fill="x", pady=(0, 14))
        ctk.CTkEntry(pr, textvariable=self._source_path,
                     placeholder_text="أو اكتب / الصق المسار هنا ...",
                     font=_font(11), fg_color=BG_INPUT, border_color=BORDER_C,
                     text_color=TEXT_PRI, height=40, justify="right",
                     corner_radius=10
                     ).pack(side="left", fill="x", expand=True, padx=(0, 8))
        OutlineButton(pr, text="تصفح", cmd=self._browse_source,
                      accent=T, height=40, icon="📁").pack(side="left")

        Divider(inner).pack(fill="x", pady=(0, 14))

        SectionLabel(inner, "مجلد الحفظ :", size=11
                     ).pack(anchor="e", pady=(0, 6))
        or_ = ctk.CTkFrame(inner, fg_color="transparent")
        or_.pack(fill="x")
        ctk.CTkEntry(or_, textvariable=self._output_dir,
                     font=_font(11), fg_color=BG_INPUT, border_color=BORDER_C,
                     text_color=TEXT_PRI, height=40, justify="right",
                     corner_radius=10
                     ).pack(side="left", fill="x", expand=True, padx=(0, 8))
        OutlineButton(or_, text="تصفح", cmd=self._browse_output,
                      accent=TEXT_SEC, height=40, icon="📂").pack(side="left")

    def _build_options_card(self):
        inner = self._card(self._content, "خيارات", "⚙️")
        row = ctk.CTkFrame(inner, fg_color=BG_INPUT, corner_radius=10,
                           border_width=1, border_color=BORDER_C)
        row.pack(fill="x")
        txt = ctk.CTkFrame(row, fg_color="transparent")
        txt.pack(side="right", padx=(0, 14), pady=12)
        ctk.CTkLabel(txt, text="تصفية ارساليات ما بعد 6 مساءً فقط",
                     font=_font(12, bold=True), text_color=TEXT_PRI,
                     anchor="e").pack(anchor="e")
        ctk.CTkLabel(txt,
                     text="يعرض فقط الارساليات الصادرة بعد الساعة 18:00",
                     font=_font(9), text_color=TEXT_DIM,
                     anchor="e").pack(anchor="e")
        ctk.CTkSwitch(row, text="", variable=self._after_6pm,
                      progress_color=G, button_color=TEXT_PRI,
                      button_hover_color=SUCCESS, width=50,
                      ).pack(side="left", padx=16, pady=12)

    def _build_actions_section(self):
        wrapper = ctk.CTkFrame(self._content, fg_color="transparent")
        wrapper.pack(fill="x", padx=20, pady=(18, 8))

        self._start_btn = GlowButton(
            wrapper, text="ابدأ المعالجة", cmd=self._run,
            color=G, hcolor=G_H, height=54, icon="▶")
        self._start_btn.pack(fill="x", pady=(0, 10))

        # Row 1: فتح | طباعة الشيتات | طباعة الاجمالى
        row1 = ctk.CTkFrame(wrapper, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 6))
        row1.columnconfigure(0, weight=1)
        row1.columnconfigure(1, weight=1)
        row1.columnconfigure(2, weight=1)

        self._open_btn = OutlineButton(row1, text="فتح الملف",
                                       cmd=self._open_output,
                                       accent=T, height=42, icon="📂")
        self._open_btn.grid(row=0, column=0, sticky="ew", padx=(0, 4))
        self._open_btn.configure(state="disabled")

        self._print_btn = OutlineButton(row1, text="طباعة الشيتات",
                                        cmd=self._open_print_dialog,
                                        accent=A, height=42, icon="🖨")
        self._print_btn.grid(row=0, column=1, sticky="ew", padx=4)
        self._print_btn.configure(state="disabled")

        self._print_ijmaly_btn = OutlineButton(
            row1, text="طباعة الاجمالى",
            cmd=self._open_ijmaly_print_dialog,
            accent=TEXT_GOLD, height=42, icon="📊")
        self._print_ijmaly_btn.grid(row=0, column=2, sticky="ew", padx=(4, 0))
        self._print_ijmaly_btn.configure(state="disabled")

        # Row 2: reset
        self._reset_btn = OutlineButton(wrapper, text="تصفير",
                                        cmd=self._reset,
                                        accent=TEXT_SEC, height=38, icon="↺")
        self._reset_btn.pack(fill="x", pady=(0, 8))

        prog_frame = ctk.CTkFrame(wrapper, fg_color="transparent")
        prog_frame.pack(fill="x", pady=(4, 0))
        self._progress = ctk.CTkProgressBar(
            prog_frame, height=5, progress_color=G,
            fg_color=BORDER_C, corner_radius=3)
        self._progress.pack(fill="x", pady=(0, 4))
        self._progress.set(0)
        self._status = StatusBadge(prog_frame)
        self._status.pack(anchor="e")

    def _build_log_panel(self):
        panel = ctk.CTkFrame(self, fg_color=BG_HDR, corner_radius=0, height=130)
        panel.pack(fill="x", side="bottom")
        panel.pack_propagate(False)

        ctk.CTkFrame(panel, fg_color=BORDER_C, height=1,
                     corner_radius=0).pack(fill="x")

        hdr = ctk.CTkFrame(panel, fg_color="transparent")
        hdr.pack(fill="x", padx=16, pady=(8, 4))

        lbl_row = ctk.CTkFrame(hdr, fg_color="transparent")
        lbl_row.pack(side="right")
        ctk.CTkLabel(lbl_row, text="◉", font=_font(9),
                     text_color=G).pack(side="right", padx=(0, 4))
        ctk.CTkLabel(lbl_row, text="سجل العمليات",
                     font=_font(10, bold=True),
                     text_color=TEXT_SEC).pack(side="right")

        ctk.CTkButton(hdr, text="مسح", font=_font(9),
                      width=52, height=22, fg_color=BG_INPUT,
                      hover_color=BG_HOVER, text_color=TEXT_DIM,
                      border_width=1, border_color=BORDER_C, corner_radius=6,
                      command=self._clear_log).pack(side="left")

        self._log_box = ctk.CTkTextbox(
            panel, font=_mono(11), fg_color="transparent",
            text_color="#C9D1D9", border_width=0, height=76, wrap="word")
        self._log_box.pack(fill="x", padx=16, pady=(0, 8))

        tags = {
            "start": ("#FCD34D", ("Consolas", 11, "bold")),
            "ok":    (SUCCESS,   None),
            "info":  (T,         None),
            "step":  ("#7A90A8", None),
            "error": (RED,       ("Consolas", 11, "bold")),
            "path":  ("#93C5FD", None),
        }
        for name, (fg, fnt) in tags.items():
            kw: dict = {"foreground": fg}
            if fnt:
                kw["font"] = fnt
            self._log_box._textbox.tag_config(name, **kw)
        self._log_box.configure(state="disabled")

    # ─────────────────────────────────────────
    # DROP ZONE
    # ─────────────────────────────────────────
    def _dz_hover(self, entering: bool):
        if entering:
            self._drop_zone.configure(border_color=T, fg_color="#0D1E35")
        else:
            if not self._source_path.get():
                self._drop_zone.configure(border_color=BORDER_C,
                                          fg_color=BG_INPUT)

    def _dz_set_file(self, path: str):
        self._source_path.set(path)
        self._drop_lbl.configure(
            text=f"✓  {Path(path).name}", text_color=SUCCESS)
        self._drop_sub.configure(text=path, text_color=TEXT_DIM)
        self._drop_zone.configure(border_color=G, fg_color="#091A10")

    def _on_drop(self, event):
        path = event.data.strip().strip("{}")
        if path.lower().endswith((".xlsx", ".xls", ".xlsm")):
            self._dz_set_file(path)
        else:
            messagebox.showwarning("تنبيه",
                "يرجى إسقاط ملف Excel فقط\n(.xlsx / .xls / .xlsm)")

    def _browse_source(self):
        p = filedialog.askopenfilename(
            title="اختر ملف Excel",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm")])
        if p:
            self._dz_set_file(p)

    def _browse_output(self):
        p = filedialog.askdirectory(title="اختر مجلد الحفظ")
        if p:
            self._output_dir.set(p)

    # ─────────────────────────────────────────
    # LOG
    # ─────────────────────────────────────────
    def _log(self, msg: str):
        m = msg.strip()
        if m.startswith("🚀"):
            tag = "start"
        elif "✅" in m or "✓" in m or "بنجاح" in m:
            tag = "ok"
        elif "❌" in m or "خطأ" in m:
            tag = "error"
        elif "📁" in m or ("   " in msg and (":\\" in m or "/" in m)):
            tag = "path"
        elif m.startswith("   ✓"):
            tag = "ok"
        elif m.startswith("   "):
            tag = "step"
        else:
            tag = "info"
        self._log_box._textbox.configure(state="normal")
        self._log_box._textbox.insert("end", msg + "\n", tag)
        self._log_box._textbox.see("end")
        self._log_box._textbox.configure(state="disabled")
        self._log_box.configure(state="disabled")

    def _clear_log(self):
        self._log_box._textbox.configure(state="normal")
        self._log_box._textbox.delete("1.0", "end")
        self._log_box._textbox.configure(state="disabled")

    # ─────────────────────────────────────────
    # RESET
    # ─────────────────────────────────────────
    def _reset(self):
        self._source_path.set("")
        self._drop_lbl.configure(text="اسحب ملف Excel وأفلته هنا",
                                  text_color=TEXT_SEC)
        self._drop_sub.configure(text=".xlsx  ·  .xls  ·  .xlsm",
                                  text_color=TEXT_DIM)
        self._drop_zone.configure(border_color=BORDER_C, fg_color=BG_INPUT)
        self._last_output   = None
        self._sector_totals = {}
        self._open_btn.configure(state="disabled")
        self._print_btn.configure(state="disabled")
        self._print_ijmaly_btn.configure(state="disabled")
        self._progress.set(0)
        self._status.set("جاهز", TEXT_DIM)
        self._clear_log()

    # ─────────────────────────────────────────
    # DIALOG LAUNCHERS
    # ─────────────────────────────────────────
    def _open_print_dialog(self):
        if not self._last_output or not Path(self._last_output).exists():
            messagebox.showwarning("تنبيه", "يرجى تشغيل المعالجة أولاً.")
            return
        PrintDialog(self, self._last_output)

    def _open_ijmaly_print_dialog(self):
        if not self._last_output or not Path(self._last_output).exists():
            messagebox.showwarning("تنبيه", "يرجى تشغيل المعالجة أولاً.")
            return
        IjmalyPrintDialog(self, self._last_output, self._sector_totals)

    # ─────────────────────────────────────────
    # RUN
    # ─────────────────────────────────────────
    def _run(self):
        path    = self._source_path.get().strip()
        out_dir = self._output_dir.get().strip() or str(DEFAULT_SAVE_DIR)

        if not path:
            messagebox.showwarning("تنبيه", "يرجى اختيار ملف Excel أولاً.")
            return
        if not os.path.exists(path):
            messagebox.showerror("خطأ", f"الملف غير موجود:\n{path}")
            return

        Path(out_dir).mkdir(parents=True, exist_ok=True)
        self._start_btn.configure(state="disabled", text="⏳  جارٍ المعالجة...")
        self._open_btn.configure(state="disabled")
        self._print_btn.configure(state="disabled")
        self._print_ijmaly_btn.configure(state="disabled")
        self._progress.set(0)
        self._status.set("جارٍ المعالجة ...", AMBER)
        self._clear_log()
        self._log("🚀 بدء المعالجة...")

        def worker():
            try:
                self.after(0, self._progress.set, 0.15)
                raw_df, filtered_df = load_and_filter(
                    path, self._after_6pm.get(),
                    lambda m: self.after(0, self._log, m))
                self.after(0, self._progress.set, 0.60)
                out, sector_totals = build_workbook(
                    raw_df, filtered_df, out_dir,
                    lambda m: self.after(0, self._log, m),
                    self._after_6pm.get())
                self.after(0, self._progress.set, 1.0)
                self.after(0, self._on_success, out, sector_totals)
            except Exception as e:
                self.after(0, self._on_error, str(e))

        threading.Thread(target=worker, daemon=True).start()

    # ─────────────────────────────────────────
    # CALLBACKS
    # ─────────────────────────────────────────
    def _on_success(self, out_path: str, sector_totals: dict):
        self._last_output   = out_path
        self._sector_totals = sector_totals
        self._start_btn.configure(state="normal", text="▶  ابدأ المعالجة")
        self._open_btn.configure(state="normal")
        self._print_btn.configure(state="normal")
        self._print_ijmaly_btn.configure(state="normal")
        self._status.set("اكتملت المعالجة بنجاح ✓", SUCCESS)
        self._log(f"📁 الملف:\n   {out_path}")
        messagebox.showinfo("تم بنجاح ✅",
                            f"تم إنشاء التقرير بنجاح!\n\n{out_path}")

    def _on_error(self, msg: str):
        self._start_btn.configure(state="normal", text="▶  ابدأ المعالجة")
        self._progress.set(0)
        self._status.set("حدث خطأ", RED)
        self._log(f"❌ خطأ: {msg}")
        messagebox.showerror("خطأ", msg)

    def _open_output(self):
        if self._last_output and os.path.exists(self._last_output):
            if sys.platform == "win32":
                os.startfile(self._last_output)     # type: ignore
            elif sys.platform == "darwin":
                subprocess.call(["open", self._last_output])
            else:
                subprocess.call(["xdg-open", self._last_output])


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────
if __name__ == "__main__":
    App().mainloop()